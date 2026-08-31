"""
Tiki Shop Crawler - CrawlFlow Python Plugin

Crawl toan bo san pham tu mot shop/category tren tiki.vn,
su dung Tiki.vn API (api.tiki.vn/v2/products) de lay du lieu.

Flow:
  1. fetch_data()  - Crawl tat ca san pham tu store/category (phan trang)
  2. process_data() - Chuan hoa du lieu
  3. export_data()  - Excel + dedup + cap nhat progress

Config:
  shop_url: str (bat buoc)  - URL cua shop/category tiki.vn, VD:
    https://tiki.vn/cua-hang/tiki-trading?t=product&cid=120473&cursor=0&category_id=8322&parent_id=2
  max_pages: int (mac dinh: 0 = khong gioi han)
  delay_ms: int (mac dinh: 500)
  page_size: int (mac dinh: 20, max: 100)
  project_id: str (tu dong truyen tu service)

KET QUA TRA VE (fetch_data tra ve JSON array, moi phan tu la 1 san pham):
  {
    "url": "https://tiki.vn/...",
    "name": "Ten san pham",
    "price": 3750,
    "old_price": 5000,
    "discount": 1250,
    "discount_rate": 25,
    "image": "https://salt.tikicdn.com/...",
    "sku": "4427223604888",
    "seller_id": 1,
    "seller_name": "Tiki Trading",
    "brand_name": "Claro",
    "brand_id": 165071,
    "category_ids": [2, 8322, 7741, 1858, 1869, 67335],
    "primary_category_name": "Bút bi bấm",
    "rating_average": 0,
    "review_count": 0,
    "availability": 1,
    "seller_product_id": 279325149,
    "seller_product_sku": "4427223604888",
    "crawled_at": "2026-08-31T12:00:00"
  }

XUAT EXCEL/CSV:
  Cot: STT, Gia bia, Gia ban, Giam gia, Phan tram, Ton kho, Don vi,
       Kich thuoc, Tinh trang, Thuong hieu, Ten san pham, Danh muc,
       Nguoi ban, So luong review
"""

import json
import hashlib
import time
import os
import re
import urllib.parse
from datetime import datetime
from html import unescape


# ── Mac dinh cho tiki.vn ────────────────────────────────────────
DEFAULT_API_BASE = "https://api.tiki.vn/v2/products"
DEFAULT_PAGE_SIZE = 20
MAX_PAGE_SIZE = 100


# ── Plugin registration ─────────────────────────────────────────
def register_presets():
    """Dang ky preset cho Tiki Shop Crawler plugin."""
    return json.dumps([
        {
            "id": "tiki-shop-crawler",
            "name": "Tiki Shop/Store Crawler",
            "description": "Crawl tat ca san pham tu mot shop hoac category tren tiki.vn",
            "icon": "ShoppingCartIcon",
            "icon_color": "#e53935",
            "project_settings": {
                "name": "",
                "description": "Crawl san pham tu tiki.vn",
                "enabled": True,
                "crawlDelay": 500,
                "userAgent": "CrawlFlow/1.0",
                "concurrency": 1,
                "executionMode": "queue",
                "groupExport": True,
                "groupFormat": "name",
                "refreshStrategy": "update_only",
                "updateMethod": "check_first_page_until_duplicate",
                "refreshInterval": 3600
            },
            "nodes": [
                {
                    "id": "ds-tiki",
                    "type": "start",
                    "position": {"x": 50, "y": 50},
                    "data": {
                        "pluginConfig": {"shop_url": ""},
                        "pluginSourceType": "tiki-shop-crawler",
                        "sourceType": "url",
                        "sourceValue": "",
                        "urlSettings": {
                            "httpClient": {
                                "clientType": "reqwest",
                                "headless": False
                            }
                        }
                    },
                    "deletable": True,
                    "draggable": True,
                    "width": 320,
                    "height": 127,
                    "zIndex": 0,
                    "parentNode": None
                },
                {
                    "id": "pre-1",
                    "type": "preprocessor",
                    "position": {"x": -568, "y": 38},
                    "data": {
                        "csvDelimiter": ",",
                        "csvHasHeader": True,
                        "extractRules": [],
                        "inputType": "json",
                        "itemSelector": "",
                        "jsonItemPath": "",
                        "urlPatterns": [],
                    },
                    "deletable": True,
                    "draggable": True,
                    "width": 320,
                    "height": 316,
                    "zIndex": 0,
                    "selected": False,
                    "dragging": False,
                },
                {
                    "id": "repository-node",
                    "type": "repository",
                    "position": {"x": 50, "y": 329},
                    "data": {},
                    "deletable": True,
                    "draggable": True,
                    "width": 320,
                    "height": 183,
                    "zIndex": 0,
                    "parentNode": None,
                },
                {
                    "id": "worker-1",
                    "type": "worker",
                    "position": {"x": 40, "y": 641},
                    "data": {
                        "detectionLogic": "and",
                        "detectionRules": [
                            {
                                "condition": "exists",
                                "id": "1783651265684",
                                "pattern": "tiki\\.vn.*-detail-\\d+",
                                "selector": "",
                                "type": "url-format",
                                "value": "",
                            }
                        ],
                    },
                    "deletable": True,
                    "draggable": True,
                    "width": 320,
                    "height": 211,
                    "zIndex": 0,
                    "parentNode": None,
                },
                {
                    "id": "ext-1",
                    "type": "html-data-extractor",
                    "position": {
                        "x": -492.44587280108254,
                        "y": 435.9208389715832,
                    },
                    "data": {
                        "customRules": [
                            {
                                "extract": "text",
                                "extractFrom": "html-element",
                                "id": "preset-tiki-html-1",
                                "name": "name",
                                "selector": "h1.product-name",
                            },
                            {
                                "extract": "text",
                                "extractFrom": "html-element",
                                "id": "preset-tiki-html-2",
                                "name": "price",
                                "selector": "[class*='price-current']",
                            },
                            {
                                "extract": "html",
                                "extractFrom": "html-element",
                                "id": "preset-tiki-html-3",
                                "name": "description",
                                "selector": "[class*='product-description']",
                            },
                            {
                                "attribute": "src",
                                "extract": "attribute",
                                "extractFrom": "html-element",
                                "id": "preset-tiki-html-4",
                                "name": "image_url",
                                "selector": "[class*='product-image'] img",
                            },
                        ],
                        "presets": ["ecommerce-product"],
                        "inspectorUrl": "https://tiki.vn/",
                        "inspectorLoading": False,
                        "inspectorHtmlContent": "",
                    },
                    "deletable": True,
                    "draggable": True,
                    "width": 320,
                    "height": 239,
                    "zIndex": 0,
                    "selected": False,
                    "dragging": False,
                    "positionAbsolute": {
                        "x": -492.44587280108254,
                        "y": 435.9208389715832,
                    },
                },
                {
                    "id": "proc-1",
                    "type": "processor",
                    "position": {"x": 52, "y": 914},
                    "data": {
                        "processorType": "generate-excel-file",
                        "settings": {
                            "autoMapHeaders": True,
                            "columnMapping": {},
                            "fileName": "tiki_crawl_results_{{date}}.xlsx",
                            "includeHeader": True,
                            "sheetName": "Sheet1",
                        },
                    },
                    "deletable": True,
                    "draggable": True,
                    "width": 320,
                    "height": 167,
                    "zIndex": 0,
                    "parentNode": None,
                },
                {
                    "id": "fetch-data-ds-tiki",
                    "type": "fetchData",
                    "position": {"x": 50, "y": 370},
                    "data": {
                        "sourceType": "url",
                        "label": "Fetch Data (url)",
                    },
                    "deletable": False,
                    "width": 320,
                    "height": 243,
                },
                {
                    "id": "completion-node",
                    "type": "completion",
                    "position": {"x": 52, "y": 1214},
                    "data": {},
                    "deletable": False,
                    "draggable": False,
                    "width": 320,
                    "height": 157,
                    "zIndex": 0,
                },
            ],
            "edges": [
                {
                    "id": "e-ds-pre",
                    "source": "ds-tiki",
                    "target": "pre-1",
                    "sourceHandle": None,
                    "targetHandle": None,
                    "type": "smoothstep",
                    "animated": False,
                    "data": {},
                },
                {
                    "id": "e-repo-worker",
                    "source": "repository-node",
                    "target": "worker-1",
                    "sourceHandle": None,
                    "targetHandle": None,
                    "type": "smoothstep",
                    "animated": False,
                    "data": {},
                },
                {
                    "id": "e-ext-worker",
                    "source": "ext-1",
                    "target": "worker-1",
                    "sourceHandle": None,
                    "targetHandle": None,
                    "type": "smoothstep",
                    "animated": False,
                    "data": {},
                },
                {
                    "id": "e-worker-proc",
                    "source": "worker-1",
                    "target": "proc-1",
                    "sourceHandle": None,
                    "targetHandle": None,
                    "type": "smoothstep",
                    "animated": False,
                    "data": {},
                },
                {
                    "id": "e-pre-1-fetch-data-ds-tiki",
                    "source": "pre-1",
                    "target": "fetch-data-ds-tiki",
                    "animated": True,
                },
                {
                    "id": "e-fetch-data-ds-tiki-repository-node",
                    "source": "fetch-data-ds-tiki",
                    "target": "repository-node",
                    "animated": True,
                },
                {
                    "id": "e-proc-1-completion-node",
                    "source": "proc-1",
                    "target": "completion-node",
                    "type": "smoothstep",
                },
            ],
        }
    ])


# ── URL parsing ─────────────────────────────────────────────────
def parse_tiki_url(url):
    """Phan tich URL tiki.vn de trích thong tin store/category.

    URL mau: https://tiki.vn/cua-hang/tiki-trading?t=product&cid=120473&cursor=0&category_id=8322&parent_id=2

    Tra ve dict voi cac truong:
      - store_slug: str (vd: 'tiki-trading')
      - category_id: str
      - parent_id: str
      - cid: str
    """
    parsed = urllib.parse.urlparse(url)
    path_parts = parsed.path.strip("/").split("/")

    # Extract store slug from path: /cua-hang/{slug}
    store_slug = ""
    if len(path_parts) >= 2 and path_parts[0] == "cua-hang":
        store_slug = path_parts[1]

    params = urllib.parse.parse_qs(parsed.query)

    return {
        "store_slug": store_slug,
        "category_id": (params.get("category_id") or [None])[0] or "",
        "parent_id": (params.get("parent_id") or [None])[0] or "",
        "cid": (params.get("cid") or [None])[0] or "",
        "cursor": (params.get("cursor") or [None])[0] or "0",
    }


def _build_api_url(shop_url, page_size, page):
    """Xay dung API URL tu shop_url."""
    parsed = parse_tiki_url(shop_url)
    params = {
        "limit": page_size,
        "page": page,
    }
    if parsed["category_id"]:
        params["category_id"] = parsed["category_id"]
    if parsed["parent_id"]:
        params["parent_id"] = parsed["parent_id"]

    query = urllib.parse.urlencode(params)
    return f"{DEFAULT_API_BASE}?{query}"


# ── Reusable "library" filter ───────────────────────────────────
def tiki_filter_parsed_data(data_json):
    """Filter chay tu dong tren parsed data cua tung item.

    Chuan hoa image URL, loai bo cache bloat fields.
    """
    try:
        items = json.loads(data_json) if isinstance(data_json, str) else data_json
    except Exception:
        return data_json

    for item in items:
        if not isinstance(item, dict):
            continue
        if "image" in item:
            item["image"] = _upgrade_tiki_image(item.get("image", ""))
        if "image_url" in item:
            item["image_url"] = _upgrade_tiki_image(item.get("image_url", ""))

    return json.dumps(items)


def _upgrade_tiki_image(url):
    """Chuyen URL anh tiki.vn tu kich thuoc nho sang kich thuoc lon.

    Mau: https://salt.tikicdn.com/cache/280x280/ts/product/...
    ->  https://salt.tikicdn.com/cache/w1280/ts/product/...
    """
    if not url:
        return url
    url = url.strip()
    # Upgrade cache size from /cache/{w}x{h}/ to /cache/w1280/
    url = re.sub(r'/cache/\d+x\d+/', '/cache/w1280/', url)
    return url


# ── HTML parser helper ──────────────────────────────────────────
class HTMLContentParser:
    """Parser HTML don gian de lay text content."""

    def __init__(self):
        self.text_parts = []
        self._capture = False
        self._depth = 0
        self._target_tag = None
        self._target_attrs = None

    def extract_text(self, html):
        self.text_parts = []
        self.feed(html)
        return " ".join(self.text_parts).strip()

    def handle_data(self, data):
        if self._capture:
            stripped = data.strip()
            if stripped:
                self.text_parts.append(stripped)

    def handle_starttag(self, tag, attrs):
        if self._target_tag and tag == self._target_tag:
            if self._target_attrs:
                attr_dict = dict(attrs)
                for k, v in self._target_attrs.items():
                    if attr_dict.get(k) == v:
                        self._capture = True
                        self._depth += 1
                        break
            else:
                self._capture = True
                self._depth += 1
        elif self._capture:
            self._depth += 1

    def handle_endtag(self, tag):
        if self._capture:
            self._depth -= 1
            if self._depth <= 0:
                self._capture = False
                self._target_tag = None
                self._target_attrs = None


def _get_text(html, selector):
    """Don gian: lay text content tu HTML."""
    p = HTMLContentParser()
    return p.extract_text(html)


# ── Helpers ─────────────────────────────────────────────────────
def _safe_float(val, default=0):
    try:
        cleaned = re.sub(r'[^\d.,]', '', str(val))
        cleaned = cleaned.replace('.', '').replace(',', '.')
        return float(cleaned) if cleaned else default
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0):
    try:
        return int(float(re.sub(r'[^\d]', '', str(val)))) if re.sub(r'[^\d]', '', str(val)) else default
    except (ValueError, TypeError):
        return default


def _clean_category(category_ids):
    """Chuyen category_ids thanh chuoi danh muc."""
    if not category_ids:
        return ""
    if isinstance(category_ids, list):
        return ";".join(str(c) for c in category_ids if c)
    return str(category_ids)


def _format_availability(availability):
    """Chuyen trang thai availability thanh text."""
    if availability == 1:
        return "Con hang"
    elif availability == 0:
        return "Het hang"
    elif availability == 2:
        return "Tram het"
    elif availability == 3:
        return "Ngoi hang"
    else:
        return str(availability)


# ── Core crawl logic ────────────────────────────────────────────
def _crawl_all_products(shop_url, max_pages, page_size, delay_ms, client_type=None, headless=None, project_id=None, db_path=None, refresh_strategy="refresh", update_method="check_first_page_until_duplicate"):
    """Tu crawl toan bo san pham cua store/category tren tiki.vn."""

    def _fetch(url):
        return crawlflow.fetch_url(url, None, client_type, headless)

    products = []
    seen_urls = set()
    page_num = 1
    max_retries = 3
    retry_base_delay = 2.0

    done_pages = set()
    if project_id:
        try:
            done_pages = set(crawlflow.get_done_pages(project_id))
        except Exception:
            done_pages = set()

    while True:
        if max_pages and page_num > max_pages:
            break
        if page_num in done_pages:
            crawlflow.log(
                f"[TikiShop][preprocess] Bo qua page {page_num} (da done, resume)",
                "info",
            )
            page_num += 1
            continue

        crawlflow.log(
            f"[TikiShop][fetch_data] Fetching page {page_num}...",
            "info",
        )

        api_url = _build_api_url(shop_url, page_size, page_num)
        crawlflow.log(f"[TikiShop][fetch_data] API URL: {api_url}", "info")

        listing_fetched = False
        for attempt in range(max_retries):
            try:
                raw = _fetch(api_url)
                listing_result = json.loads(raw) if isinstance(raw, str) else raw
                listing_fetched = True
                break
            except Exception as e:
                if attempt < max_retries - 1:
                    delay = retry_base_delay * (2 ** attempt)
                    crawlflow.log(
                        f"[TikiShop] Loi fetch page {page_num}, retry {attempt + 2}/{max_retries} sau {delay}s — {e}",
                        "warn",
                    )
                    time.sleep(delay)
                else:
                    crawlflow.log(
                        f"[TikiShop] Loi fetch page {page_num} sau {max_retries} lan: {e}",
                        "error",
                    )
        if not listing_fetched:
            break

        if not isinstance(listing_result, dict):
            crawlflow.log(f"[TikiShop] Ket qua tra ve khong phai dict: {type(listing_result)}", "error")
            break

        data_list = listing_result.get("data", [])
        paging = listing_result.get("paging", {})

        if not data_list:
            crawlflow.log(f"[TikiShop] Trang {page_num} khong co san pham. Ket thuc.", "info")
            break

        crawlflow.log(
            f"[TikiShop] Tim thay {len(data_list)} san pham o trang {page_num} / {paging.get('last_page', '?')}",
            "info",
        )

        # Luu URL vao DB de progress bar cap nhat realtime
        if data_list and project_id and db_path:
            raw_items = []
            for item in data_list:
                p_url = item.get("url_path", "") or item.get("url_key", "")
                if p_url:
                    if not p_url.startswith("http"):
                        p_url = "https://tiki.vn/" + p_url.lstrip("/")
                    raw_items.append({
                        "source_url": p_url,
                        "item_type": "url",
                        "item_hash": hashlib.sha256(p_url.encode("utf-8")).hexdigest(),
                    })
            try:
                res = json.loads(crawlflow.save_raw_items(project_id, db_path, json.dumps(raw_items)))
                saved = int(res.get("inserted", 0))
                crawlflow.log(
                    f"[TikiShop] Da luu {saved} URL moi vao DB (trang {page_num})",
                    "info",
                )
            except Exception as e:
                crawlflow.log(f"[TikiShop] Loi save_raw_items: {e}", "warn")

        if project_id:
            try:
                crawlflow.mark_page_done(project_id, api_url, page_num, len(data_list))
            except Exception:
                pass

        # Parse each product in this page
        for item in data_list:
            prod = _parse_product_from_api(item)
            products.append(prod)
            name = prod.get("name") or "(khong ten)"
            price = prod.get("price") or ""
            crawlflow.log(
                f"[TikiShop] OK — \"{name}\" | gia: {price} | page {page_num}",
                "info",
            )

            # Emit live progress
            if project_id:
                try:
                    crawlflow.emit_event(project_id, "progress", json.dumps({
                        "items_total": len(data_list),
                        "items_processed": len(products),
                        "items_success": len(products),
                        "items_failed": 0,
                        "items_pending": len(data_list) - len(products),
                        "progress_pct": round(len(products) / len(data_list) * 100, 1) if data_list else 0.0,
                        "phase": "fetching",
                        "message": f"[TikiShop] OK — {name}",
                        "last_run_at": "",
                    }))
                except Exception:
                    pass

        # Check pagination
        if max_pages and page_num >= max_pages:
            crawlflow.log(f"[TikiShop] Da du max_pages={max_pages}", "info")
            break

        total_pages = paging.get("last_page", 0)
        if page_num >= total_pages:
            crawlflow.log(f"[TikiShop] Het phan trang tai trang {page_num}/{total_pages}", "info")
            break

        page_num += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000.0)

    crawlflow.log(
        f"[TikiShop] Hoan tat: {len(products)} san pham",
        "info",
    )
    return products


def _parse_product_from_api(item):
    """Parse 1 item tu Tiki.vn API thanh dict chuan."""
    return {
        "url": item.get("url_path", "") or item.get("url_key", ""),
        "name": item.get("name", ""),
        "price": _safe_float(item.get("price", 0)),
        "old_price": _safe_float(item.get("original_price", 0)),
        "discount": _safe_float(item.get("discount", 0)),
        "discount_rate": _safe_float(item.get("discount_rate", 0)),
        "image": item.get("thumbnail_url", ""),
        "sku": item.get("sku", ""),
        "seller_id": item.get("seller_id", ""),
        "seller_name": item.get("seller_name", ""),
        "brand_id": item.get("brand_id", ""),
        "brand_name": item.get("brand_name", ""),
        "category_ids": item.get("category_ids", []),
        "primary_category_name": item.get("primary_category_name", ""),
        "rating_average": item.get("rating_average", 0),
        "review_count": item.get("review_count", 0),
        "availability": _format_availability(item.get("availability", 1)),
        "seller_product_id": item.get("seller_product_id", ""),
        "seller_product_sku": item.get("seller_product_sku", ""),
        "productset_id": item.get("productset_id", ""),
        "url_key": item.get("url_key", ""),
        "crawled_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
    }


# ── Main entry points ──────────────────────────────────────────
def fetch_data(config_json):
    """Crawl toan bo san pham cua store/category tiki.vn.

    Config:
        shop_url (str, bat buoc): URL cua store/category tiki.vn
        max_pages (int, mac dinh: 0): so trang toi da (0 = khong gioi han)
        page_size (int, mac dinh: 20): so san pham moi trang (max 100)
        delay_ms (int, mac dinh: 500): nghi giua cac request
        project_id (str): ID project
    """
    config = json.loads(config_json) if isinstance(config_json, str) else config_json

    shop_url = (config.get("shop_url") or "").strip()
    if not shop_url:
        shop_url = (config.get("source_value") or "").strip()
    if not shop_url:
        crawlflow.log("[TikiShop][fetch_data] Thieu shop_url trong config", "error")
        return json.dumps([])

    max_pages = int((config.get("max_pages") or 0) or 0)
    if max_pages < 1:
        max_pages = 0
    page_size = int((config.get("page_size") or DEFAULT_PAGE_SIZE) or DEFAULT_PAGE_SIZE)
    if page_size < 1:
        page_size = DEFAULT_PAGE_SIZE
    if page_size > MAX_PAGE_SIZE:
        page_size = MAX_PAGE_SIZE
    delay_ms = int((config.get("delay_ms") or 500) or 500)
    if delay_ms < 0:
        delay_ms = 0

    client_type = (config.get("clientType") or config.get("client_type")
                   or (config.get("urlSettings") or {}).get("httpClient", {}).get("clientType"))
    if client_type not in ("reqwest", "chrome", "cdp"):
        client_type = "reqwest"
    headless = bool((config.get("headless")
                      or (config.get("urlSettings") or {}).get("httpClient", {}).get("headless", True)))

    refresh_strategy = config.get("refresh_strategy") or "refresh"
    update_method = config.get("update_method") or "check_first_page_until_duplicate"

    crawlflow.log(
        f"[TikiShop][fetch_data] Bat dau crawl shop={shop_url} (max_pages={max_pages}, page_size={page_size})",
        "info",
    )

    products = _crawl_all_products(
        shop_url, max_pages, page_size, delay_ms,
        client_type, headless,
        config.get("project_id"), config.get("db_path"),
        refresh_strategy, update_method
    )

    # Format for Rust pipeline: item_type='url' with raw_content
    items = []
    for p in products:
        item_url = p.get("url", "")
        raw_html = p.pop("raw_html", "")
        items.append({
            "source_url": item_url,
            "item_type": "url",
            "item_hash": hashlib.sha256(item_url.encode("utf-8")).hexdigest() if item_url else hashlib.sha256(json.dumps(p, ensure_ascii=False).encode("utf-8")).hexdigest(),
            "raw_content": raw_html,
            "extracted_url": item_url,
        })

    crawlflow.log(
        f"[TikiShop][fetch_data] Hoan tat: {len(items)} san pham",
        "info",
    )
    return json.dumps(items)


def process_data(data_json, config_json):
    """Chuan hoa du lieu san pham tu Tiki.vn API."""
    data = json.loads(data_json) if isinstance(data_json, str) else data_json

    total = len(data)
    crawlflow.log(f"[TikiShop][process] Bat dau chuan hoa {total} san pham", "info")

    normalized = []
    for idx, item in enumerate(data, 1):
        norm = {
            "url": item.get("url", ""),
            "name": (item.get("name") or "").strip(),
            "price": _safe_float(item.get("price", 0)),
            "old_price": _safe_float(item.get("old_price", 0)),
            "discount": _safe_float(item.get("discount", 0)),
            "discount_rate": _safe_float(item.get("discount_rate", 0)),
            "image": item.get("image", ""),
            "image_url": item.get("image", ""),
            "sku": item.get("sku", "").strip(),
            "seller_id": item.get("seller_id", ""),
            "seller_name": item.get("seller_name", ""),
            "brand_id": item.get("brand_id", ""),
            "brand_name": item.get("brand_name", ""),
            "category_ids": item.get("category_ids", []),
            "primary_category_name": item.get("primary_category_name", ""),
            "rating_average": item.get("rating_average", 0),
            "review_count": item.get("review_count", 0),
            "availability": _format_availability(item.get("availability", 1)),
            "seller_product_id": item.get("seller_product_id", ""),
            "seller_product_sku": item.get("seller_product_sku", ""),
            "crawled_at": item.get("crawled_at", datetime.now().strftime("%Y-%m-%dT%H:%M:%S")),
        }
        name = norm["name"] or "(khong ten)"
        price_str = f" | gia: {norm['price']}" if norm["price"] else ""
        crawlflow.log(
            f"[TikiShop][process] [{idx}/{total}] \"{name}\"{price_str} | {norm['url']}",
            "debug",
        )
        normalized.append(norm)

    crawlflow.log(f"[TikiShop][process] Hoan tinh chuan hoa {len(normalized)}/{total} san pham", "info")
    return json.dumps(normalized)


def export_data(data_json, config_json):
    """Xuat du lieu ra Excel/CSV voi co che append + check trung."""
    data = json.loads(data_json) if isinstance(data_json, str) else data_json
    config = json.loads(config_json) if isinstance(config_json, str) else config_json

    project_id = config.get("project_id", "default")
    output_dir = config.get("output_dir")
    if not output_dir:
        output_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(output_dir, exist_ok=True)

    project_name = config.get("projectName", "")
    shop_url = config.get("shop_url", "")
    shop_name = "tiki_shop"

    if project_name:
        shop_name = project_name.replace("?", "_").replace("&", "_").replace("/", "_").replace("\\", "_")
    elif shop_url:
        parsed = parse_tiki_url(shop_url)
        shop_name = parsed.get("store_slug") or "tiki_shop"

    started_at = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    total_items = len(data)

    crawlflow.log(f"[TikiShop] Bat dau export {total_items} san pham", "info")

    # Dedup
    dedup_path = os.path.join(output_dir, f".{shop_name}_dedup.json")
    seen_ids = set()
    if os.path.exists(dedup_path):
        try:
            content = crawlflow.read_file(dedup_path)
            seen_ids = set(json.loads(content))
            crawlflow.log(f"[TikiShop] Da doc {len(seen_ids)} ID da xu ly tu file dedup", "info")
        except Exception as e:
            crawlflow.log(f"[TikiShop] Loi doc dedup file: {e}", "warn")

    new_products = []
    for item in data:
        dedup_key = item.get("url", "") or item.get("sku", "")
        if dedup_key and dedup_key in seen_ids:
            continue
        new_products.append(item)

    crawlflow.log(f"[TikiShop] Sau dedup: {len(new_products)} san pham moi (da bo qua {len(data) - len(new_products)} san pham trung)", "info")

    if not new_products:
        crawlflow.log("[TikiShop] Khong co san pham moi de export", "info")
        _update_progress(project_id, {
            "items_total": total_items,
            "items_processed": total_items,
            "items_success": 0,
            "items_failed": 0,
            "progress_pct": 100.0,
            "avg_time_ms": 0,
            "total_time_ms": 0,
            "started_at": started_at,
            "message": "Khong co san pham moi",
        })
        return json.dumps({"file": "", "count": 0, "new": 0, "skipped": len(data)})

    excel_path = os.path.join(output_dir, f"{shop_name}_products.xlsx")
    csv_path = os.path.join(output_dir, f"{shop_name}_products.csv")

    try:
        from openpyxl import Workbook, load_workbook
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False

    if has_openpyxl:
        count = _export_xlsx(new_products, excel_path, seen_ids)
        crawlflow.log(f"[TikiShop] Da ghi {count} san pham vao {excel_path}", "info")
    else:
        count = _export_csv(new_products, csv_path)
        crawlflow.log(f"[TikiShop] Da ghi {count} san pham vao {csv_path} (CSV)", "info")

    for item in new_products:
        dedup_key = item.get("url", "") or item.get("sku", "")
        if dedup_key:
            seen_ids.add(dedup_key)

    try:
        crawlflow.save_file(dedup_path, json.dumps(list(seen_ids), ensure_ascii=False))
    except Exception as e:
        crawlflow.log(f"[TikiShop] Loi ghi dedup file: {e}", "warn")

    elapsed = (datetime.now() - datetime.strptime(started_at, "%Y-%m-%dT%H:%M:%S")).total_seconds() * 1000
    _update_progress(project_id, {
        "items_total": total_items,
        "items_processed": total_items,
        "items_success": count,
        "items_failed": total_items - count,
        "progress_pct": 100.0,
        "avg_time_ms": elapsed / max(total_items, 1),
        "total_time_ms": elapsed,
        "started_at": started_at,
        "message": f"Export hoan thanh: {count} san pham moi",
    })

    output_file = excel_path if has_openpyxl else csv_path
    result = {
        "file": output_file,
        "count": count,
        "new": len(new_products),
        "skipped": len(data) - len(new_products),
    }
    return json.dumps(result)


def _export_xlsx(products, filepath, seen_ids):
    """Ghi san pham vao file Excel (append neu file da ton tai)."""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        crawlflow.log(f"[TikiShop] Mo file Excel co san: {filepath}", "info")
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "San pham"
        headers = [
            "STT", "Gia bia", "Gia ban", "Giam gia", "Phan tram",
            "Ton kho", "Don vi", "Kich thuoc", "Tinh trang",
            "Thuong hieu", "Ten san pham", "Danh muc", "Nguoi ban",
            "So luong review", "SKU", "URL"
        ]
        ws.append(headers)
        crawlflow.log(f"[TikiShop] Tao file Excel moi: {filepath}", "info")

    count = 0
    for item in products:
        dedup_key = item.get("url", "") or item.get("sku", "")
        if dedup_key and dedup_key in seen_ids:
            continue

        specs = item.get("specs", {})
        row = [
            ws.max_row,
            item.get("old_price", 0),
            item.get("price", 0),
            item.get("discount", 0),
            item.get("discount_rate", 0),
            item.get("stock", ""),
            item.get("unit", ""),
            _spec_value(specs, "Kich thuoc", "Size", "Kich co"),
            item.get("availability", ""),
            item.get("brand_name", ""),
            item.get("name") or item.get("product_name", ""),
            item.get("primary_category_name", ""),
            item.get("seller_name", ""),
            item.get("review_count", 0),
            item.get("sku", ""),
            item.get("url", ""),
        ]
        ws.append(row)
        count += 1

    wb.save(filepath)
    crawlflow.log(f"[TikiShop] Da them {count} dong vao Excel", "info")
    return count


def _export_csv(products, filepath):
    """Fallback: ghi CSV."""
    import csv

    mode = "a" if os.path.exists(filepath) else "w"
    has_header = mode == "w"

    existing_count = 0
    if mode == "a" and os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8-sig") as f:
            existing_count = sum(1 for _ in f) - 1

    with open(filepath, mode, newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if has_header:
            writer.writerow([
                "STT", "Gia bia", "Gia ban", "Giam gia", "Phan tram",
                "Ton kho", "Don vi", "Kich thuoc", "Tinh trang",
                "Thuong hieu", "Ten san pham", "Danh muc", "Nguoi ban",
                "So luong review", "SKU", "URL"
            ])

        count = existing_count
        for item in products:
            writer.writerow([
                count + 1,
                item.get("old_price", 0),
                item.get("price", 0),
                item.get("discount", 0),
                item.get("discount_rate", 0),
                item.get("stock", ""),
                item.get("unit", ""),
                _spec_value(item.get("specs", {}), "Kich thuoc", "Size"),
                item.get("availability", ""),
                item.get("brand_name", ""),
                item.get("name") or item.get("product_name", ""),
                item.get("primary_category_name", ""),
                item.get("seller_name", ""),
                item.get("review_count", 0),
                item.get("sku", ""),
                item.get("url", ""),
            ])
            count += 1

    crawlflow.log(f"[TikiShop] Da them {count} dong vao CSV", "info")
    return count


def _spec_value(specs, *keys):
    """Lay gia tri tu specs bang nhieu key co the co."""
    for key in keys:
        val = specs.get(key)
        if val:
            return str(val).strip()
    return ""


def _update_progress(project_id, data):
    """Cap nhat progress vao Rust backend."""
    try:
        def _to_u64(v):
            try:
                return int(round(float(v))) if v is not None else 0
            except (TypeError, ValueError):
                return 0

        info = {
            "items_total": _to_u64(data.get("items_total", 0)),
            "items_processed": _to_u64(data.get("items_processed", 0)),
            "items_success": _to_u64(data.get("items_success", 0)),
            "items_failed": _to_u64(data.get("items_failed", 0)),
            "progress_pct": float(data.get("progress_pct", 0.0) or 0.0),
            "avg_time_ms": float(data.get("avg_time_ms", 0.0) or 0.0),
            "total_time_ms": _to_u64(data.get("total_time_ms", 0)),
            "started_at": data.get("started_at", ""),
            "message": data.get("message", ""),
        }
        crawlflow.update_progress(project_id, json.dumps(info))
    except Exception as e:
        crawlflow.log(f"[TikiShop] Loi update progress: {e}", "error")


# ── Plugin hooks ────────────────────────────────────────────────
def register_preprocessors():
    """Dang ky preprocessor cho Tiki Shop Crawler."""
    return json.dumps([{
        "id": "tiki-shop-products",
        "name": "Tiki Shop Products",
        "plugin_id": "",
        "input_type": "json",
        "platform": "tiki.vn",
        "config": {
            "input_type": "json",
            "item_selector": None,
            "url_patterns": [],
            "extract_rules": [],
            "csv_delimiter": None,
            "csv_has_header": None,
            "json_item_path": None,
            "client_type": None,
            "client_timeout_secs": None,
            "client_headless": None,
            "wait_for_selector": None,
            "wait_for_content": None,
            "wait_timeout_ms": None,
        },
    }])


def on_load(config=None):
    crawlflow.log("[TikiShop] Plugin loaded", "info")
    try:
        import openpyxl
        crawlflow.log("[TikiShop] openpyxl available - will use Excel output", "info")
    except ImportError:
        crawlflow.log("[TikiShop] openpyxl not installed - will use CSV output", "warn")

    try:
        crawlflow.register_filter("parsed_data", tiki_filter_parsed_data)
        crawlflow.log("[TikiShop] Registered 'parsed_data' filter", "info")
    except Exception as e:
        crawlflow.log(f"[TikiShop] register_filter failed: {e}", "warn")


def on_unload():
    crawlflow.log("[TikiShop] Plugin unloaded", "info")