"""
Oreka Shop Crawler - CrawlFlow Python Plugin

Crawl toan bo san pham tu mot shop tren oreka.vn,
parse du lieu va xuat ra file Excel voi co che append + check trung.

Flow:
  1. fetch_data()  - Crawl tat ca san pham tu shop (phan trang)
  2. process_data() - Chuan hoa du lieu
  3. export_data()  - Excel + dedup + cap nhat progress

Config:
  shop_url: str (bat buoc)  - URL cua shop, VD: https://oreka.vn/shop/ten-shop
  max_pages: int (mặc định: 0 = không giới hạn)
  delay_ms: int (mặc định: 1500)
  selectors: dict (tuỳ chon) - Ghi de selector mac dinh
  output_dir: str (mặc định: thu muc hien tai)
  project_id: str (tu dong truyen tu service)

KET QUA TRA VE (fetch_data tra ve JSON array, moi phan tu la 1 san pham):
  {
    "url": "https://oreka.vn/product/...",
    "name": "Ten san pham",
    "price": 299000,
    "old_price": 399000,
    "image": "https://oreka.vn/.../img.jpg",
    "sku": "SP001",
    "description": "Mo ta san pham...",
    "specs": { "Thuong hieu": "...", "Chat lieu": "..." },
    "category": "Danh muc",
    "availability": "Con hang",
    "stock": "1",
    "crawled_at": "2026-07-02T12:00:00"
  }

XUAT EXCEL/CSV:
  Cot: STT, Gia bia, Gia ban, Ton kho, Don vi,
       Khoi luong (g), Kich thuoc, Tinh trang, Nam XB,
       Thuong hieu, Ten sach, Danh muc
"""

import json
import hashlib
import time
import os
import re
import urllib.parse
from datetime import datetime
from html import unescape
from html.parser import HTMLParser


def register_presets():
    """Register preset for Oreka Shop Crawler plugin."""
    preset = {
        "id": "oreka-shop-crawler",
        "name": "Oreka Shop Crawler",
        "description": "Crawl sản phẩm từ shop oreka.vn với custom extraction rules và export ra Excel",
        "icon": "ShoppingCartIcon",
        "icon_color": "#10b981",
        "project_settings": {
            "name": "",
            "description": "Crawl sản phẩm từ shop oreka.vn",
            "enabled": True,
            "crawlDelay": 1500,
            "userAgent": "CrawlFlow/1.0",
            "concurrency": 1,
            "executionMode": "queue",
            "groupExport": True,
            "groupFormat": "name",
            "refreshStrategy": "update_only",
            "updateMethod": "check_first_page_until_duplicate",
            "refreshInterval": 3600
        },
        "nodes": [
            {
                "id": "ds-oreka",
                "type": "start",
                "position": {"x": 50, "y": 50},
                "data": {
                    "pluginConfig": {"shop_url": ""},
                    "pluginSourceType": "oreka-shop-crawler",
                    "sourceType": "url",
                    "sourceValue": "",
                    "urlSettings": {
                        "httpClient": {
                            "clientType": "reqwest",
                            "headless": False
                        }
                    }
                },
                "deletable": True,
                "draggable": True,
                "width": 320,
                "height": 127,
                "zIndex": 0,
                "parentNode": None
            },
            {
                "id": "pre-1",
                "type": "preprocessor",
                "position": {"x": -568, "y": 38},
                "data": {
                    "csvDelimiter": ",",
                    "csvHasHeader": True,
                    "extractRules": [],
                    "inputType": "html",
                    "itemSelector": ".mt-12.grid.grid-cols-5.gap-10",
                    "jsonItemPath": "",
                    "urlPatterns": [
                        {
                            "enabled": True,
                            "type": "regex",
                            "value": ".*-detail\\/[0-9]{1,}\\/?",
                        }
                    ],
                },
                "deletable": True,
                "draggable": True,
                "width": 320,
                "height": 316,
                "zIndex": 0,
                "selected": False,
                "dragging": False,
            },
            {
                "id": "repository-node",
                "type": "repository",
                "position": {"x": 50, "y": 329},
                "data": {},
                "deletable": True,
                "draggable": True,
                "width": 320,
                "height": 183,
                "zIndex": 0,
                "parentNode": None,
            },
            {
                "id": "worker-1",
                "type": "worker",
                "position": {"x": 40, "y": 641},
                "data": {
                    "detectionLogic": "and",
                    "detectionRules": [
                        {
                            "condition": "exists",
                            "id": "1783651265684",
                            "pattern": ".*-detail\\/[0-9]{1,}\\/?",
                            "selector": "",
                            "type": "url-format",
                            "value": "",
                        }
                    ],
                },
                "deletable": True,
                "draggable": True,
                "width": 320,
                "height": 211,
                "zIndex": 0,
                "parentNode": None,
            },
            {
                "id": "ext-1",
                "type": "html-data-extractor",
                "position": {
                    "x": -492.44587280108254,
                    "y": 435.9208389715832,
                },
                "data": {
                    "customRules": [
                        {
                            "extract": "text",
                            "extractFrom": "html-element",
                            "id": "preset-ecom-html-1",
                            "name": "product_name",
                            "selector": "h1",
                        },
                        {
                            "extract": "text",
                            "extractFrom": "html-element",
                            "id": "preset-ecom-html-2",
                            "name": "price",
                            "selector": "p.font-semibold.text-16.leading-8.text-black-600.line-clamp-1.break-all.styles_productPrice__zkPlt",
                        },
                        {
                            "extract": "text",
                            "extractFrom": "json-ld",
                            "id": "preset-ecom-html-3",
                            "jsonPath": "offers.sku",
                            "name": "sku",
                        },
                        {
                            "extract": "html",
                            "extractFrom": "html-element",
                            "id": "preset-ecom-html-4",
                            "name": "description",
                            "selector": "div.mt-6.whitespace-pre-wrap > p.text",
                        },
                        {
                            "attribute": "src",
                            "extract": "attribute",
                            "extractFrom": "html-element",
                            "id": "preset-ecom-html-5",
                            "name": "image_url",
                            "selector": "img.styles_imageThumb__OYuNp.object-cover.rounded-md",
                        },
                        {
                            "attribute": "src",
                            "extract": "attribute",
                            "extractFrom": "html-element",
                            "extractMultiple": True,
                            "id": "preset-ecom-html-6",
                            "name": "images",
                            "selector": ".image-gallery-thumbnail img",
                        },
                        {
                            "id": "1785642132986",
                            "name": "category",
                            "extractFrom": "html-element",
                            "selector": (
                                "#__next > div.bg-\\[\\#F7F7F7\\]:nth-of-type(4)"
                                " > div.bg-black-f3:nth-of-type(5)"
                                " > div.container.mx-auto.mb-12.pb-20.styles_mainContainer__VzpQg"
                                " > div.flex.styles_mainContainerSeller__7DVE5:nth-of-type(3)"
                                " > div.flex-1.styles_aboutSeller__kHT9J:nth-of-type(1)"
                                " > div.styles_containerOutstanding__CHgHx:nth-of-type(2)"
                                " > div > div.flex.styles_itemInfo__FFfqL:nth-of-type(1)"
                                " > div.styles_customContainerBc__j_y_v"
                                " > ul.flex.container.mx-auto.py-4.flex-wrap.styles_customUl__Z_Nd_ li a"
                            ),
                            "extract": "text",
                            "extractMultiple": True,
                        },
                    ],
                    "presets": ["ecommerce-product"],
                    "inspectorUrl": "https://www.oreka.vn/mua-ban-sach-thieu-nhi/bo-sach-nhung-cam-xuc-nho-quan-trong-cua-be--6-cuon----bia-cung--in-mau-detail/1088773",
                    "inspectorLoading": False,
                    "inspectorHtmlContent": "",
                },
                "deletable": True,
                "draggable": True,
                "width": 320,
                "height": 239,
                "zIndex": 0,
                "selected": False,
                "dragging": False,
                "positionAbsolute": {
                    "x": -492.44587280108254,
                    "y": 435.9208389715832,
                },
            },
            {
                "id": "proc-1",
                "type": "processor",
                "position": {"x": 52, "y": 914},
                "data": {
                    "processorType": "generate-excel-file",
                    "settings": {
                        "autoMapHeaders": True,
                        "columnMapping": {},
                        "fileName": "crawl_results_{{date}}.xlsx",
                        "includeHeader": True,
                        "sheetName": "Sheet1",
                    },
                },
                "deletable": True,
                "draggable": True,
                "width": 320,
                "height": 167,
                "zIndex": 0,
                "parentNode": None,
            },
            {
                "id": "fetch-data-ds-oreka",
                "type": "fetchData",
                "position": {"x": 50, "y": 370},
                "data": {
                    "sourceType": "url",
                    "label": "Fetch Data (url)",
                },
                "deletable": False,
                "width": 320,
                "height": 243,
            },
            {
                "id": "completion-node",
                "type": "completion",
                "position": {"x": 52, "y": 1214},
                "data": {},
                "deletable": False,
                "draggable": False,
                "width": 320,
                "height": 157,
            },
        ],
        "edges": [
            {
                "id": "e-ds-pre",
                "source": "ds-oreka",
                "target": "pre-1",
                "sourceHandle": None,
                "targetHandle": None,
                "type": "smoothstep",
                "animated": False,
                "data": {},
            },
            {
                "id": "e-repo-worker",
                "source": "repository-node",
                "target": "worker-1",
                "sourceHandle": None,
                "targetHandle": None,
                "type": "smoothstep",
                "animated": False,
                "data": {},
            },
            {
                "id": "e-ext-worker",
                "source": "ext-1",
                "target": "worker-1",
                "sourceHandle": None,
                "targetHandle": None,
                "type": "smoothstep",
                "animated": False,
                "data": {},
            },
            {
                "id": "e-worker-proc",
                "source": "worker-1",
                "target": "proc-1",
                "sourceHandle": None,
                "targetHandle": None,
                "type": "smoothstep",
                "animated": False,
                "data": {},
            },
            {
                "id": "e-pre-1-fetch-data-ds-oreka",
                "source": "pre-1",
                "target": "fetch-data-ds-oreka",
                "animated": True,
            },
            {
                "id": "e-fetch-data-ds-oreka-repository-node",
                "source": "fetch-data-ds-oreka",
                "target": "repository-node",
                "animated": True,
            },
            {
                "id": "e-proc-1-completion-node",
                "source": "proc-1",
                "target": "completion-node",
                "type": "smoothstep",
            },
        ],
    }
    return json.dumps([preset])


def _add_page_to_url(url, page_param, page_num):
    parsed = urllib.parse.urlparse(url)
    qd = urllib.parse.parse_qs(parsed.query)
    qd[page_param] = [str(page_num)]
    new_query = urllib.parse.urlencode(qd, doseq=True)
    return urllib.parse.urlunparse((
        parsed.scheme,
        parsed.netloc,
        parsed.path,
        parsed.params,
        new_query,
        parsed.fragment
    ))


def find_store_id(data):
    if isinstance(data, dict):
        if "store" in data and isinstance(data["store"], dict) and "id" in data["store"]:
            return data["store"]["id"]
        if "storeId" in data:
            return data["storeId"]
        for val in data.values():
            res = find_store_id(val)
            if res:
                return res
    elif isinstance(data, list):
        for val in data:
            res = find_store_id(val)
            if res:
                return res
    return None


def _find_apollo_store_id(data):
    """Lay store ID tu __APOLLO_STATE__ cua Oreka Next.js page."""
    if isinstance(data, dict):
        apollo_state = data.get("__APOLLO_STATE__")
        if isinstance(apollo_state, dict):
            for cache_key, record in apollo_state.items():
                if not cache_key.startswith("Store:"):
                    continue
                store_id = record.get("id") if isinstance(record, dict) else None
                return str(store_id or cache_key.removeprefix("Store:")).strip()
        for value in data.values():
            store_id = _find_apollo_store_id(value)
            if store_id:
                return store_id
    elif isinstance(data, list):
        for value in data:
            store_id = _find_apollo_store_id(value)
            if store_id:
                return store_id
    return None


def _extract_store_slug_from_url(url):
    """Extract store slug from Oreka store URL."""
    # Pattern: https://www.oreka.vn/store/SLUG or /store/SLUG
    match = re.search(r'/store/([A-Za-z0-9\-\.]+)', url)
    if match:
        return match.group(1)
    return None


def _extract_store_id_from_next_data(html):
    """Parse JSON trong script __NEXT_DATA__ va lay Store:<id> tu Apollo cache."""
    match = re.search(
        r'<script\b(?=[^>]*\bid=["\']__NEXT_DATA__["\'])[^>]*>(.*?)</script>',
        html,
        re.DOTALL | re.IGNORECASE,
    )
    if not match:
        return None

    try:
        next_data = json.loads(unescape(match.group(1)).strip())
    except (TypeError, ValueError, json.JSONDecodeError):
        return None

    apollo_state = (
        next_data.get("props", {})
        .get("pageProps", {})
        .get("__APOLLO_STATE__", {})
    )
    if not isinstance(apollo_state, dict):
        return _find_apollo_store_id(next_data)

    for cache_key, record in apollo_state.items():
        if cache_key.startswith("Store:"):
            store_id = record.get("id") if isinstance(record, dict) else None
            return str(store_id or cache_key.removeprefix("Store:")).strip()
    return None


def _extract_store_id_from_html(html):
    """Lay store ID tu HTML trang Oreka, uu tien du lieu Next.js."""
    html = unescape(html)
    store_id = _extract_store_id_from_next_data(html)
    if store_id:
        return store_id

    # Direct fallback for Apollo Store UUID pattern (e.g. Store:950fc091-0766-4b5e-af6a-ad7b5325a1fb)
    uuid_match = re.search(r'["\']Store:([a-fA-F0-9\-]{36})["\']', html)
    if uuid_match:
        return uuid_match.group(1).strip()

    # Fallback to look inside social meta og:image URL pattern (e.g. store-950fc091-0766-4b5e-af6a-ad7b5325a1fb.webp)
    meta_match = re.search(r'store-([a-fA-F0-9\-]{36})\.', html)
    if meta_match:
        return meta_match.group(1).strip()

    for script_content in re.findall(r'<script\b[^>]*>(.*?)</script>', html, re.DOTALL | re.IGNORECASE):
        if "storeId" not in script_content and '"store"' not in script_content:
            continue
        try:
            script_data = json.loads(script_content.strip())
            store_id = _find_apollo_store_id(script_data) or find_store_id(script_data)
            if store_id:
                return str(store_id).strip()
        except (TypeError, ValueError, json.JSONDecodeError):
            continue

    for pattern in (
        r'["\']storeId["\']\s*:\s*["\']([^"\']+)["\']',
        r'["\']store["\']\s*:\s*\{\s*["\']id["\']\s*:\s*["\']([^"\']+)["\']',
        r'[?&]storeId=([^&"\'\s]+)',
    ):
        match = re.search(pattern, html, re.IGNORECASE)
        if match:
            return urllib.parse.unquote(match.group(1)).strip()
    return None


def _oreka_listing_url(source_url, store_id):
    parsed = urllib.parse.urlparse(source_url)
    base_url = f"{parsed.scheme or 'https'}://{parsed.netloc or 'www.oreka.vn'}"
    query = urllib.parse.urlencode({
        'storeId': store_id,
        'sort': 'createdAt',
        'order': 'desc',
    })
    return f"{base_url}/mua-ban?{query}"


def register_preprocessors():
    """Dang ky preprocessor chuyen trang store Oreka thanh danh sach san pham."""
    return json.dumps([{
        "id": "oreka-store-products",
        "name": "Oreka Store Products",
        "plugin_id": "",
        "input_type": "html",
        "platform": "oreka.vn",
        "config": {
            "input_type": "html",
            "item_selector": None,
            "url_patterns": [],
            "extract_rules": [],
            "csv_delimiter": None,
            "csv_has_header": None,
            "json_item_path": None,
            "client_type": None,
            "client_timeout_secs": None,
            "client_headless": None,
            "wait_for_selector": None,
            "wait_for_content": None,
            "wait_timeout_ms": None,
        },
    }])


def preprocess_data(data_json):
    """Lay store ID tu HTML nguon, fetch tung trang listing va luu product URL
    vao DB NGAY khi tim thay (de progress bar cap nhat realtime).

    Khac voi truoc day (tra ve 1 loat 'listing_url' de Rust fetch sau), o day
    plugin tu fetch tung trang, trich product URL va goi crawlflow.save_raw_items
    luu tung luot vao DB. UI ('items_pending') se tang len tung buoc thay vi
    chi nhay 1 lan sau khi xong het.

    Tra ve [] de bao cho Rust rang item da duoc luu truc tiep vao DB
    (Rust se skip Stage B trich URL trung lap).
    """
    payload = json.loads(data_json) if isinstance(data_json, str) else data_json
    html = payload.get("raw_data", "")
    source_url = payload.get("source_url", "https://www.oreka.vn")
    project_id = payload.get("project_id", "")
    db_path = payload.get("db_path", "")
    config = payload.get("config", {})
    refresh_strategy = payload.get("refresh_strategy", "refresh")

    # Skip listing crawl for update_only — existing items in DB are sufficient.
    # Only refresh / refresh_update need to re-crawl listing pages.
    if refresh_strategy == "update_only":
        crawlflow.log("[OrekaShop][preprocess] update_only — skip listing crawl", "info")
        return json.dumps([])

    store_id = _extract_store_id_from_html(html) or _extract_store_id_from_html(source_url)

    if not store_id:
        crawlflow.log("[OrekaShop][preprocess] Khong tim thay storeId trong HTML", "warn")
        return json.dumps([])

    base_listing_url = _oreka_listing_url(source_url, store_id)

    # Pagination: Oreka store listing supports ?page=N.
    # max_pages = 0 hoac None => unlimited.
    max_pages = int((config.get("max_pages") or 0) or 0)
    if max_pages < 1:
        max_pages = 0  # 0 = unlimited

    # Cac page da hoan thanh (tu cycle truoc) se duoc skip khi resume.
    done_pages = set()
    if project_id:
        try:
            done_pages = set(crawlflow.get_done_pages(project_id))
        except Exception:
            done_pages = set()

    total_saved = 0
    page_num = 1
    delay_ms = int((config.get("delay_ms") or 1000) or 1000)

    while True:
        if max_pages and page_num > max_pages:
            break
        page_url = _add_page_to_url(base_listing_url, "page", page_num)

        if page_num in done_pages:
            crawlflow.log(
                f"[OrekaShop][preprocess] Bo qua page {page_num} (da done, resume)",
                "info",
            )
            page_num += 1
            continue

        crawlflow.log(
            f"[OrekaShop][preprocess] Fetch listing page {page_num}: {page_url}",
            "info",
        )

        try:
            raw = crawlflow.fetch_url(page_url, None, "reqwest", False)
            result = json.loads(raw) if isinstance(raw, str) else raw
            listing_html = result.get("body", "") if isinstance(result, dict) else ""
        except Exception as e:
            crawlflow.log(
                f"[OrekaShop][preprocess] Loi fetch page {page_num}: {e}", "error"
            )
            break

        # Detect redirect: chi break neu thuc su redirect ve trang 1 khi dang o trang > 1
        final_url = result.get("final_url", "") if isinstance(result, dict) else ""
        if final_url and final_url != page_url:
            parsed_final = urllib.parse.urlparse(final_url)
            parsed_original = urllib.parse.urlparse(page_url)
            final_params = urllib.parse.parse_qs(parsed_final.query)
            original_params = urllib.parse.parse_qs(parsed_original.query)
            
            # Lay page number tu final_url (mac dinh la 1 neu khong co page parameter)
            final_page = int(final_params.get("page", ["1"])[0])
            original_page = int(original_params.get("page", ["1"])[0])
            
            # Chi break neu redirect tu trang > 1 ve trang 1 (het trang)
            # Khong break chi vi parameter doi thu tu
            if original_page > 1 and final_page == 1:
                crawlflow.log(f"[OrekaShop][preprocess] Phat hien redirect tu trang {original_page} ve trang 1 (het trang)", "warn")
                crawlflow.log(f"[OrekaShop][preprocess] Dung tai trang {page_num}", "info")
                break

        if not listing_html:
            crawlflow.log(
                f"[OrekaShop][preprocess] Listing page {page_num} rong", "warn"
            )
            break

        product_urls = _extract_oreka_listing_links(listing_html, page_url)
        crawlflow.log(
            f"[OrekaShop][preprocess] Tim thay {len(product_urls)} product URL o trang {page_num}",
            "info",
        )

        # Luu tung luot product URL vao DB NGAY de UI cap nhat realtime.
        if product_urls and project_id and db_path:
            raw_items = []
            for p_url in product_urls:
                raw_items.append({
                    "source_url": p_url,
                    "item_type": "url",
                    "item_hash": hashlib.sha256(p_url.encode("utf-8")).hexdigest(),
                })
            try:
                res = json.loads(crawlflow.save_raw_items(project_id, db_path, json.dumps(raw_items)))
                saved = int(res.get("inserted", 0))
                total_saved += saved
                crawlflow.log(
                    f"[OrekaShop][preprocess] Da luu {saved} URL moi (tong {total_saved}) vao DB",
                    "info",
                )
            except Exception as e:
                crawlflow.log(
                    f"[OrekaShop][preprocess] Loi save_raw_items: {e}", "warn"
                )

        # Danh dau page done de resume.
        if project_id:
            try:
                crawlflow.mark_page_done(project_id, page_url, page_num, len(product_urls))
            except Exception:
                pass

        has_next = _has_next_page(listing_html, page_num)
        if not has_next:
            crawlflow.log(
                f"[OrekaShop][preprocess] Het phan trang tai trang {page_num}", "info"
            )
            break

        page_num += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000.0)

    crawlflow.log(
        f"[OrekaShop][preprocess] Hoan tat: {total_saved} product URL da luu vao DB "
        f"(max_pages={max_pages or 'unlimited'}, bo qua {len(done_pages)} done)",
        "info",
    )
    return json.dumps([])


# ── Mac dinh cho oreka.vn ──────────────────────────────────────────────
DEFAULT_SELECTORS = {
    "product_list": ".product-list, .products-grid, [class*='product-grid'], [class*='list-product'], .shop-products",
    "product_item": "a[href*='/product/'], a[href*='/san-pham/'], .product-item a, [class*='product-card'] a, .item-product a",
    "product_link_attr": "href",
    "pagination": ".pagination, .page-list, [class*='pagination'], .pages",
    "next_page": "a.next, a[rel='next'], .pagination a:last-child, a:contains('Sau')",
    "page_param": "page",
}

DEFAULT_DETAIL = {
    "name": "h1.product-name, h1.product-title, [class*='product-name'], [class*='product-title'], .product_detail_name, h1",
    "price": "[class*='price'] [class*='current'], [class*='product-price'] [class*='special'], .product_price .price, [class*='product-price'] > span, .current-price, span.price",
    "old_price": "[class*='old-price'], [class*='price-old'], .product_price .old-price, del.price, .old-price",
    "image": "meta[property='og:image'], .product-gallery img, [class*='product-image'] img, .product_detail_img img, img.main-image",
    "sku": "[class*='sku'], .product-code, [class*='product-id'], .product_sku, span.sku",
    "description": "[class*='description'], .product-description, .product_detail_description, #product-description, .tab-content",
    "availability": "[class*='availability'], .product-status, .stock, span:contains('Con hang'), span:contains('Het hang')",
    "category": "[class*='breadcrumb'] a:last-child, .breadcrumbs a:last-child, [class*='breadcrumb'] li:last-child a, .breadcrumb li:last-child",
    "specs_table": "table. specifications, .product-attributes table, .product-specs table, [class*='specs'] table, .parameter table",
}


class HTMLContentParser(HTMLParser):
    """Parser HTML don gian de lay text va attribute."""

    def __init__(self):
        super().__init__()
        self.text_parts = []
        self._capture = False
        self._depth = 0
        self._target_tag = None
        self._target_attrs = None
        self.result_attrs = {}

    def extract_text(self, html):
        self.text_parts = []
        self.feed(html)
        return " ".join(self.text_parts).strip()

    def handle_data(self, data):
        if self._capture:
            stripped = data.strip()
            if stripped:
                self.text_parts.append(stripped)

    def handle_starttag(self, tag, attrs):
        if self._target_tag and tag == self._target_tag:
            if self._target_attrs:
                attr_dict = dict(attrs)
                for k, v in self._target_attrs.items():
                    if attr_dict.get(k) == v:
                        self._capture = True
                        self._depth += 1
                        break
            else:
                self._capture = True
                self._depth += 1
        elif self._capture:
            self._depth += 1

    def handle_endtag(self, tag):
        if self._capture:
            self._depth -= 1
            if self._depth <= 0:
                self._capture = False
                self._target_tag = None
                self._target_attrs = None


def _get_text(html, selector):
    """Don gian: lay text content tu HTML."""
    p = HTMLContentParser()
    return p.extract_text(html)


def _spec_value(specs, *keys):
    """Lay gia tri tu specs bang nhieu key co the co."""
    for key in keys:
        val = specs.get(key)
        if val:
            return str(val).strip()
    return ""


def _safe_float(val, default=0):
    try:
        cleaned = re.sub(r'[^\d.,]', '', str(val))
        cleaned = cleaned.replace('.', '').replace(',', '.')
        return float(cleaned) if cleaned else default
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0):
    try:
        return int(float(re.sub(r'[^\d]', '', str(val)))) if re.sub(r'[^\d]', '', str(val)) else default
    except (ValueError, TypeError):
        return default


def _clean_category(value):
    """Chuan hoa chuoi category:
    - Loai bo dau '>' (va cac ky tu phan tach breadcrumb nhu ; | / › >>)
    - Bo 2 items dau tien (thuong la 'Oreka', 'Đồ cũ')
    - Noi cac phan con lai bang ';'
    """
    if value is None:
        return ""
    # Dua ve danh sach cac doan
    if isinstance(value, list):
        parts = [str(v) for v in value]
    else:
        parts = re.split(r'\s*(?:>|>>|›|»|\||;)\s*', str(value))
    # Lam sach tung doan (bo dau >/khoang trang o dau va cuoi, loai bo o trong)
    cleaned = []
    for p in parts:
        q = re.sub(r'^[\s>›»]+', '', str(p))
        q = re.sub(r'[\s>›»]+$', '', q)
        q = q.replace('>', '').replace('›', '').replace('»', '').strip()
        if q:
            cleaned.append(q)
    # Bo 2 phan dau tien
    if len(cleaned) > 2:
        cleaned = cleaned[2:]
    return ";".join(cleaned)


# Kich thuoc anh lon nhat duoc ho tro boi static.oreka.vn CDN
_OREKA_IMAGE_SIZE = "800-800"

# Regex bat tien to kich thuoc dang "250-250_", "800-800_", ...
_OREKA_SIZE_PREFIX_RE = re.compile(r'^(https?://[^/]*static\.oreka\.vn/)\d+-\d+_(.*)$')

# Cac URL khong phai anh san pham thuc su (anh quang cao / thumbnail cua oreka)
_OREKA_AD_IMAGE_RE = re.compile(r'static\.oreka\.vn/(?:[^/]+/)*_next/static/', re.IGNORECASE)


def _upgrade_oreka_image(url):
    """Chuyen URL anh oreka.vn tu kich thuoc nho (vd 250-250) sang kich thuoc lon (800-800).

    Chi ap dung voi CDN static.oreka.vn co tien to <w>-<h>_.
    Cac URL quang cao / thumbnail cua oreka (vd /d/_next/static/...thumbnail...)
    bi loai bo (tra ve chuoi rong).
    """
    if not url:
        return url
    url = url.strip()
    # Loai bo anh quang cao / thumbnail cua oreka
    if _OREKA_AD_IMAGE_RE.search(url):
        return ""
    m = _OREKA_SIZE_PREFIX_RE.match(url)
    if not m:
        return url
    return f"{m.group(1)}{_OREKA_IMAGE_SIZE}_{m.group(2)}"


def _upgrade_oreka_images(value):
    """Upgrade mot hoac nhieu URL anh (chuoi cach nhau boi khoang trang hoac list).

    Cac URL bi loai bo se khong xuat hien trong ket qua.
    """
    if value is None:
        return value
    if isinstance(value, list):
        return [u for u in (_upgrade_oreka_image(v) for v in value) if u]
    if isinstance(value, str):
        parts = value.split()
        upgraded = [_upgrade_oreka_image(u) for u in parts]
        kept = [u for u in upgraded if u]
        if len(parts) > 1:
            return " ".join(kept)
        return kept[0] if kept else ""
    return value


# ── Reusable "library" filter, registered with the backend ───────────────
# Rust invokes this automatically on every item's parsed data (the `images`
# array) — no hard-coded field surgery inside process_data.
def oreka_filter_parsed_data(data_json):
    """Filter chay tu dong tren parsed data cua tung item.

    Input: JSON string cua mot list gom 1 object item.
    Output: JSON string cua list (cung kich thuoc) voi image/images da duoc
    chuan hoa (len size 800-800, loai bo anh quang cao oreka).
    """
    try:
        items = json.loads(data_json) if isinstance(data_json, str) else data_json
    except Exception:
        return data_json

    for item in items:
        if not isinstance(item, dict):
            continue
        if "image" in item:
            item["image"] = _upgrade_oreka_image(item.get("image", ""))
        if "image_url" in item:
            item["image_url"] = _upgrade_oreka_image(item.get("image_url", ""))
        if "images" in item:
            item["images"] = _upgrade_oreka_images(item.get("images"))

    return json.dumps(items)


def _extract_attr(html, attr="src"):
    """Lay attribute value tu the HTML dau tien."""
    m = re.search(rf'{attr}\s*=\s*["\']([^"\']+)["\']', html)
    return m.group(1) if m else ""


def _extract_meta_content(html):
    """Lay content tu the meta tag."""
    m = re.search(r'content\s*=\s*["\']([^"\']+)["\']', html)
    return m.group(1) if m else ""


def _extract_specs(html):
    """Lay bang thong so san pham."""
    specs = {}
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL)
    for row in rows:
        cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row, re.DOTALL)
        if len(cells) >= 2:
            key = _get_text(cells[0], "").strip()
            val = _get_text(cells[1], "").strip()
            if key:
                specs[key] = val
    return specs


def _extract_listing_links(html, base_url):
    """Trich xuat tat ca link san pham tu trang danh sach."""
    links = set()
    for m in re.finditer(r'<a[^>]*href\s*=\s*["\'](/[^"\']*(?:product|san-pham)[^"\']*)["\']', html, re.IGNORECASE):
        href = m.group(1)
        full = href if href.startswith("http") else (base_url.rstrip("/") + "/" + href.lstrip("/"))
        links.add(full)
    for m in re.finditer(r'<a[^>]*href\s*=\s*["\'](https?://[^"\']*(?:product|san-pham)[^"\']*)["\']', html, re.IGNORECASE):
        links.add(m.group(1))
    return list(links)


def _extract_oreka_listing_links(html, base_url):
    """Lay URL chi tiet san pham Oreka.

    Uu tien 1: Lay tu JSON-LD @type=ItemList (chinh xac nhat, Oreka nhung day).
    Uu tien 2: Lay tu the <a href> co pattern /mua-ban-*/...-detail/<id>.
    """
    import json as _json
    links = set()
    origin = "/".join(base_url.rstrip("/").split("/")[:3])  # https://www.oreka.vn

    # --- Uu tien 1: JSON-LD ItemList ---
    for ld_text in re.findall(
        r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        html, re.DOTALL | re.IGNORECASE
    ):
        try:
            data = _json.loads(ld_text)
            # Ho tro ca object don va @graph array
            nodes = []
            if isinstance(data, list):
                nodes = data
            elif isinstance(data, dict):
                nodes = data.get("@graph", [data])
            for node in nodes:
                if not isinstance(node, dict):
                    continue
                if node.get("@type") == "ItemList":
                    for item in node.get("itemListElement", []):
                        url = item.get("url", "")
                        if url:
                            full = urllib.parse.urljoin(origin + "/", url.lstrip("/"))
                            links.add(full)
        except Exception:
            pass

    if links:
        return links

    # --- Uu tien 2: <a href> regex ---
    # Pattern: /mua-ban-<category>/<slug>-detail/<id>
    pattern = re.compile(
        r'href=["\']((https?://[^"\'/][^"\']*)?' +
        r'/mua-ban(?:-[^/"\' ]+)?/[^"\' ]*-detail/\d+(?:\?[^"\' ]*)?)["\']',
        re.IGNORECASE,
    )
    for match in pattern.finditer(unescape(html)):
        href = match.group(1)
        links.add(urllib.parse.urljoin(origin + "/", href.lstrip("/")))

    return links


def _has_next_page(html, current_page):
    """Kiem tra xem trang hien tai co link/nut sang trang tiep theo khong.

    Chien luoc (theo thu tu uu tien):
    1. rel="next"  — chuan HTML, chinh xac nhat.
    2. Ton tai href chua ?page=<N+1> hoac &page=<N+1>  — chinh xac voi Oreka.
    3. aria-label chua "next" / "sau" / "tiep".
    4. class chua "next" / "forward" / "after" tren the <a> hoac <button>.
    5. Text cua link la ky tu phan trang pho bien: ›, », Next, Sau, Tiep.
    """
    next_page_num = current_page + 1

    # 0. Tim tat ca so trang tu link ?page=N / &page=N trong HTML. Neu co
    #    trang lon hon current -> chac chan con trang sau (Oreka render day du
    #    pagination nen cach nay rat tin cay).
    page_nums = [
        int(m)
        for m in re.findall(r'[?&]page=(\d+)', html)
        if m.isdigit()
    ]
    if page_nums and max(page_nums) > current_page:
        return True

    # 1. rel="next"
    if re.search(r'rel=["\']next["\']', html, re.IGNORECASE):
        return True

    # 2. Href co page=<N+1> — Oreka render link nay trong SSR HTML
    if re.search(
        r'href=["\'][^"\']*[?&]page=' + str(next_page_num) + r'(?:[&"\'/]|$)',
        html, re.IGNORECASE
    ):
        return True

    # 2b. Oreka specific: nut next la <a> chua <img alt="arrow-right">
    #     href cua nut nay luon tro den trang tiep theo
    if re.search(
        r'<a\b[^>]*href=[^>]+>\s*(?:<[^>]+>\s*)*<img[^>]+alt=["\']arrow-right["\']',
        html, re.IGNORECASE | re.DOTALL
    ):
        return True

    # 3. aria-label next/sau/tiep
    if re.search(
        r'aria-label=["\'][^"\']*(next|sau|tiep theo|trang sau)[^"\']*(?:["\']|$)',
        html, re.IGNORECASE
    ):
        return True

    # 4. class next/forward/after tren <a> hoac <button>
    if re.search(
        r'<(?:a|button)[^>]*class=["\'][^"\']*(\bnext\b|\bforward\b|\bafter\b)[^"\']*["\']',
        html, re.IGNORECASE
    ):
        return True

    # 5. Text pho bien cua nut next (›, », Next, Sau, Tiep)
    next_text_pattern = re.compile(
        r'<(?:a|button|li)[^>]*>[^<]{0,40}(?:›|»|&rsaquo;|&raquo;|\bNext\b|\bSau\b|\bTiep\b)[^<]{0,10}</(?:a|button|li)>',
        re.IGNORECASE
    )
    if next_text_pattern.search(html):
        return True

    return False


def _parse_product_from_html(html, url):
    """Phan tich HTML trang chi tiet de lay thong tin san pham."""
    parser = HTMLContentParser()
    text = parser.extract_text(html)

    name = ""
    price = 0
    old_price = 0
    image = ""
    sku = ""
    description = ""
    availability = ""
    category = ""
    specs = {}

    # Ten san pham - uu tien the h1
    h1_match = re.search(r'<h1[^>]*>(.*?)</h1>', html, re.DOTALL)
    if h1_match:
        name = _get_text(h1_match.group(1), "").strip()

    # Fallback: extract name from JSON-LD BreadcrumbList (last item = product)
    if not name:
        for ld_text in re.findall(
            r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html, re.DOTALL | re.IGNORECASE
        ):
            try:
                ld_data = json.loads(ld_text)
                if isinstance(ld_data, dict) and ld_data.get("@type") == "BreadcrumbList":
                    items = ld_data.get("itemListElement", [])
                    if items:
                        last_item = items[-1]
                        name = last_item.get("item", {}).get("name", "").strip()
                        if name:
                            break
            except (json.JSONDecodeError, AttributeError):
                pass

    # Fallback: extract name from URL slug
    # Pattern: /mua-ban-<category>/<slug>-detail/<id>
    if not name:
        slug_match = re.search(r'/mua-ban[^/]*/([^/]+)-detail/', url)
        if slug_match:
            name = slug_match.group(1).replace('-', ' ').strip()

    # Gia
    price_patterns = [
        r'<span[^>]*class\s*=\s*["\'][^"\']*price[^"\']*["\'][^>]*>([^<]+)</span>',
        r'<div[^>]*class\s*=\s*["\'][^"\']*price[^"\']*["\'][^>]*>([^<]+)</div>',
        r'<ins[^>]*>([^<]+)</ins>',
        r'<span[^>]*class\s*=\s*["\'][^"\']*current[^"\']*["\'][^>]*>([^<]+)</span>',
    ]
    for pat in price_patterns:
        m = re.search(pat, html)
        if m:
            price = _safe_float(m.group(1))
            if price > 0:
                break

    if price == 0:
        gia_matches = re.findall(r'(?:(\d[\d.,]*)\s*(?:₫|vnđ|vnd|đ))', html, re.IGNORECASE)
        if gia_matches:
            price = _safe_float(gia_matches[0])

    # Gia cu
    old_patterns = [
        r'<span[^>]*class\s*=\s*["\'][^"\']*old[^"\']*["\'][^>]*>([^<]+)</span>',
        r'<del[^>]*>([^<]+)</del>',
        r'<strike[^>]*>([^<]+)</strike>',
    ]
    for pat in old_patterns:
        m = re.search(pat, html)
        if m:
            old_price = _safe_float(m.group(1))
            if old_price > 0:
                break
    if old_price == 0 and price > 0:
        old_price = price

    # Hinh anh - uu tien og:image
    og_image = re.search(r'<meta[^>]*property\s*=\s*["\']og:image["\'][^>]*content\s*=\s*["\']([^"\']+)["\']', html)
    if og_image:
        image = og_image.group(1)
    else:
        img_match = re.search(r'<img[^>]*class\s*=\s*["\'][^"\']*(?:product|main|gallery)[^"\']*["\'][^>]*src\s*=\s*["\']([^"\']+)["\']', html)
        if img_match:
            image = img_match.group(1)
        else:
            img_match = re.search(r'<img[^>]*src\s*=\s*["\']([^"\']+(?:product|san-pham)[^"\']+)["\']', html)
            if img_match:
                image = img_match.group(1)

    # SKU
    sku_matches = re.findall(r'(?:SKU|Mã sản phẩm|Product Code|Mã SP)\s*[:;]\s*([^\s<]+)', html, re.IGNORECASE)
    if sku_matches:
        sku = sku_matches[0].strip()

    # Mo ta - lay tu tab content hoac div mo ta
    desc_match = re.search(r'<div[^>]*(?:description|product-desc|tab-content|product_description)[^>]*>(.*?)</div>\s*</div>', html, re.DOTALL)
    if desc_match:
        description = _get_text(desc_match.group(1), "").strip()
    if not description:
        desc_match = re.search(r'<div[^>]*(?:description|product-desc)[^>]*>(.*?)</div>', html, re.DOTALL)
        if desc_match:
            description = _get_text(desc_match.group(1), "").strip()
    if not description and len(description) > 50000:
        description = description[:50000]

    # Tinh trang
    avail_match = re.search(r'(?:Còn hàng|Tạm hết|Hết hàng|Liên hệ|In stock|Out of stock)', html, re.IGNORECASE)
    if avail_match:
        availability = avail_match.group(0)

    # Ton kho
    stock = ""
    stock_match = re.search(r'(?:Số lượng|Tồn kho|Kho hàng|Còn lại)\s*[:;]?\s*(\d+)', html, re.IGNORECASE)
    if stock_match:
        stock = stock_match.group(1)
    if not stock:
        stock_match = re.search(r'(?:Còn|còn)\s*(\d+)\s*(?:sản phẩm|sp|hàng)', html, re.IGNORECASE)
        if stock_match:
            stock = stock_match.group(1)

    # Danh muc
    breadcrumb = re.search(r'<ul[^>]*class\s*=\s*["\'][^"\']*breadcrumb[^"\']*["\']>(.*?)</ul>', html, re.DOTALL)
    if breadcrumb:
        cats = re.findall(r'<a[^>]*>(.*?)</a>', breadcrumb.group(1))
        if len(cats) >= 2:
            category = _get_text(cats[-1], "").strip()

    # Bang thong so
    spec_tables = re.findall(r'<table[^>]*>(.*?)</table>', html, re.DOTALL)
    for table_html in spec_tables:
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table_html, re.DOTALL)
        for row in rows[:30]:
            cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row, re.DOTALL)
            if len(cells) >= 2:
                k = _get_text(cells[0], "").strip()
                v = _get_text(cells[1], "").strip()
                if k:
                    specs[k] = v

    return {
        "url": url,
        "name": name or os.path.basename(url),
        "price": price,
        "old_price": old_price,
        "image_url": image,
        "sku": sku,
        "description": description[:500] if description else "",
        "specs": specs,
        "category": _clean_category(category),
        "stock": stock,
        "availability": availability or "Còn hàng",
        "crawled_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "raw_html": html,
    }


def on_load(config=None):
    crawlflow.log("[OrekaShop] Plugin loaded", "info")
    try:
        import openpyxl
        crawlflow.log("[OrekaShop] openpyxl available - will use Excel output", "info")
    except ImportError:
        crawlflow.log("[OrekaShop] openpyxl not installed - will use CSV output", "warn")

    # Dang ky filter "library" chay tu dong tren parsed data cua tung item.
    # Rust se goi oreka_filter_parsed_data() moi khi co parsed data (mang images).
    try:
        crawlflow.register_filter("parsed_data", oreka_filter_parsed_data)
        crawlflow.log("[OrekaShop] Registered 'parsed_data' filter", "info")
    except Exception as e:
        crawlflow.log(f"[OrekaShop] register_filter failed: {e}", "warn")


def _crawl_all_products(shop_url, max_pages, delay_ms, client_type=None, headless=None, project_id=None, db_path=None, refresh_strategy="refresh", update_method="check_first_page_until_duplicate"):
    """Tu crawl toan bo san pham cua shop (phan trang + trich product URL + parse).

    refresh_strategy:
        'refresh'           — crawl lai hoan toan tu trang 1
        'refresh_update'    — crawl lai hoan toan (pipeline skip re-process)
        'update_only'       — chi quet data moi, dung theo update_method
    update_method (cho update_only):
        'check_first_page_until_duplicate' — quet tu trang 1, dung khi phat hien duplicate
        'check_last_page'                  — quet tu trang cuoi cung, dung khi het san pham
    """
    def _fetch(url):
        # Goi Python SDK fetch_url voi kem client_type de chon kenh reqwest/chrome.
        return crawlflow.fetch_url(url, None, client_type, headless)

    # 1. Lay storeId tu store page.
    try:
        raw = _fetch(shop_url)
        store_result = json.loads(raw) if isinstance(raw, str) else raw
    except Exception as e:
        crawlflow.log(f"[OrekaShop] Loi fetch store page: {e}", "error")
        return []

    store_html = store_result.get("body", "") if isinstance(store_result, dict) else ""
    if not store_html:
        crawlflow.log("[OrekaShop] Store page rong", "error")
        return []

    store_id = _extract_store_id_from_html(store_html) or _extract_store_id_from_html(shop_url)
    if not store_id:
        crawlflow.log("[OrekaShop] Khong tim thay storeId", "error")
        return []

    base_listing_url = _oreka_listing_url(shop_url, store_id)
    crawlflow.log(
        f"[OrekaShop] storeId={store_id} | listing base={base_listing_url}", "info"
    )

    products = []
    seen_urls = set()
    page_num = 1

    # Cac page da hoan thanh o chu ky crawl truoc (luu trong bang crawl_pages)
    # se duoc bo qua de ho tro resume khi service bi dung dot ngot.
    done_pages = set()
    if project_id:
        try:
            done_pages = set(crawlflow.get_done_pages(project_id))
        except Exception:
            done_pages = set()

    while True:
        page_url = _add_page_to_url(base_listing_url, "page", page_num)
        if page_num in done_pages:
            crawlflow.log(f"[OrekaShop] Bo qua page {page_num} (da done, resume)", "info")
            page_num += 1
            if max_pages and page_num > max_pages:
                break
            continue
        crawlflow.log(f"[OrekaShop] Listing page {page_num}: {page_url}", "info")

        try:
            raw = _fetch(page_url)
            listing_result = json.loads(raw) if isinstance(raw, str) else raw
        except Exception as e:
            crawlflow.log(f"[OrekaShop] Loi fetch listing page {page_num}: {e}", "error")
            break

        # Detect redirect: chi break neu thuc su redirect ve trang 1 khi dang o trang > 1
        final_url = listing_result.get("final_url", "") if isinstance(listing_result, dict) else ""
        if final_url and final_url != page_url:
            parsed_final = urllib.parse.urlparse(final_url)
            parsed_original = urllib.parse.urlparse(page_url)
            final_params = urllib.parse.parse_qs(parsed_final.query)
            original_params = urllib.parse.parse_qs(parsed_original.query)
            
            # Lay page number tu final_url (mac dinh la 1 neu khong co page parameter)
            final_page = int(final_params.get("page", ["1"])[0])
            original_page = int(original_params.get("page", ["1"])[0])
            
            # Chi break neu redirect tu trang > 1 ve trang 1 (het trang)
            # Khong break chi vi parameter doi thu tu
            if original_page > 1 and final_page == 1:
                crawlflow.log(f"[OrekaShop] Phat hien redirect tu trang {original_page} ve trang 1 (het trang)", "warn")
                crawlflow.log(f"[OrekaShop] Dung tai trang {page_num}", "info")
                break

        listing_html = listing_result.get("body", "") if isinstance(listing_result, dict) else ""
        if not listing_html:
            crawlflow.log(f"[OrekaShop] Listing page {page_num} rong", "warn")
            break

        # Trich product URLs tu listing HTML.
        product_urls = _extract_oreka_listing_links(listing_html, page_url)
        crawlflow.log(
            f"[OrekaShop] Tim thay {len(product_urls)} product URL o trang {page_num}",
            "info",
        )

        # Neu khong tim thay URL nao => het san pham
        if not product_urls and refresh_strategy != "refresh":
            crawlflow.log(f"[OrekaShop] Khong con san pham, dung tai trang {page_num}", "info")
            break

        # Luu tung luot product URL vao DB NGAY de progress bar (pending)
        # cap nhat realtime thay vi chi nhay 1 lan sau khi xong het.
        saved = 0
        if product_urls and project_id and db_path:
            raw_items = []
            for p_url in product_urls:
                raw_items.append({
                    "source_url": p_url,
                    "item_type": "url",
                    "item_hash": hashlib.sha256(p_url.encode("utf-8")).hexdigest(),
                })
            try:
                res = json.loads(crawlflow.save_raw_items(project_id, db_path, json.dumps(raw_items)))
                saved = int(res.get("inserted", 0))
                crawlflow.log(
                    f"[OrekaShop] Da luu {saved} URL moi vao DB (trang {page_num})",
                    "info",
                )
            except Exception as e:
                crawlflow.log(f"[OrekaShop] Loi save_raw_items: {e}", "warn")

        # Voi chien luoc update_only: dung neu khong co URL moi nao duoc insert
        if refresh_strategy == "update_only" and saved == 0 and product_urls:
            crawlflow.log(f"[OrekaShop] Khong con URL moi, dung tai trang {page_num}", "info")
            break

        # Danh dau page nay da hoan thanh de ho tro resume.
        if project_id:
            try:
                crawlflow.mark_page_done(project_id, page_url, page_num, len(product_urls))
            except Exception as e:
                crawlflow.log(f"[OrekaShop] mark_page_done loi: {e}", "warn")

        for p_url in product_urls:
            if p_url in seen_urls:
                crawlflow.log(f"[OrekaShop] Bo qua URL trung lap: {p_url}", "debug")
                continue
            seen_urls.add(p_url)
            crawlflow.log(f"[OrekaShop] Fetch product ({len(seen_urls)}/{len(product_urls)} trang {page_num}): {p_url}", "info")
            t0 = time.time()
            try:
                praw = _fetch(p_url)
                pres = json.loads(praw) if isinstance(praw, str) else praw
                phtml = pres.get("body", "") if isinstance(pres, dict) else ""
                elapsed_ms = int((time.time() - t0) * 1000)
                if phtml:
                    prod = _parse_product_from_html(phtml, p_url)
                    products.append(prod)
                    name = prod.get("name") or prod.get("title") or "(khong ten)"
                    price = prod.get("price") or prod.get("current_price") or ""
                    crawlflow.log(
                        f"[OrekaShop] OK {elapsed_ms}ms — \"{name}\"" + (f" | gia: {price}" if price else "") + f" | {p_url}",
                        "info",
                    )
                else:
                    crawlflow.log(f"[OrekaShop] Canh bao: HTML rong sau {elapsed_ms}ms — {p_url}", "warn")
            except Exception as e:
                elapsed_ms = int((time.time() - t0) * 1000)
                crawlflow.log(f"[OrekaShop] Loi fetch product sau {elapsed_ms}ms — {p_url} — {e}", "warn")

        # Tiep tuc phan trang: reqwest da lay xong HTML listing, giao cho
        # Python tim nut "next page" trong chinh HTML do (khong fetch them).
        if max_pages and page_num >= max_pages:
            crawlflow.log(f"[OrekaShop] Da du max_pages={max_pages}", "info")
            break

        has_next = _has_next_page(listing_html, page_num)
        if not has_next:
            crawlflow.log(f"[OrekaShop] Het phan trang tai trang {page_num}", "info")
            break
        page_num += 1
        if delay_ms > 0:
            time.sleep(delay_ms / 1000.0)

    return products


def fetch_data(config_json):
    """Crawl toan bo san pham cua shop va tra ve truc tiep N item 'product'.

    Plugin tu quyet dinh toan bo logic: lay storeId, phan trang listing,
    trich product URL, parse chi tiet. Rust chi can gom cac item nay vao
    raw_items (item_type='product') de worker + exporter xu ly tiep.

    Config:
        shop_url (str, bat buoc): URL cua store (vd: https://www.oreka.vn/store/motsach)
        max_pages (int, mac dinh: 0): gioi han so trang (0 = khong gioi han)
        delay_ms (int, mac dinh: 1000): nghi giua cac request
        project_id (str): ID project (tu dong inject)
    """
    config = json.loads(config_json) if isinstance(config_json, str) else config_json

    shop_url = (config.get("shop_url") or "").strip()
    if not shop_url:
        # Fallback: dung chinh sourceValue neu shop_url chua duoc set.
        shop_url = (config.get("source_value") or "").strip()
    if not shop_url:
        crawlflow.log("[OrekaShop] Thieu shop_url trong config", "error")
        return json.dumps([])

    max_pages = int((config.get("max_pages") or 0) or 0)
    if max_pages < 1:
        max_pages = 0  # 0 = unlimited
    delay_ms = int((config.get("delay_ms") or 1000) or 1000)
    if delay_ms < 0:
        delay_ms = 0

    # Chon kenh fetch: reqwest (mac dinh) hoac chrome. Doc tu config
    # (clientType) hoac urlSettings.httpClient.clientType cua node.
    client_type = (config.get("clientType") or config.get("client_type")
                   or (config.get("urlSettings") or {}).get("httpClient", {}).get("clientType"))
    if client_type not in ("reqwest", "chrome", "cdp"):
        client_type = "reqwest"
    headless = bool((config.get("headless")
                     or (config.get("urlSettings") or {}).get("httpClient", {}).get("headless", True)))

    crawlflow.log(
        f"[OrekaShop][fetch_data] Bat dau crawl shop={shop_url} (max_pages={max_pages}, client={client_type})",
        "info",
    )

    refresh_strategy = config.get("refresh_strategy") or "refresh"
    update_method = config.get("update_method") or "check_first_page_until_duplicate"
    crawlflow.log(
        f"[OrekaShop][fetch_data] refresh_strategy={refresh_strategy}, update_method={update_method}",
        "info",
    )

    products = _crawl_all_products(shop_url, max_pages, delay_ms, client_type, headless, config.get("project_id"), config.get("db_path"), refresh_strategy, update_method)

    # Dung dinh dang item ma Rust/worker hieu: item_type='url'.
    items = []
    for p in products:
        item_url = p.get("url", "")
        raw_html = p.pop("raw_html", "")
        items.append({
            "source_url": item_url,
            "item_type": "url",
            "item_hash": hashlib.sha256(item_url.encode("utf-8")).hexdigest() if item_url else hashlib.sha256(json.dumps(p, ensure_ascii=False).encode("utf-8")).hexdigest(),
            "raw_content": raw_html,
            "extracted_url": item_url,
        })

    crawlflow.log(
        f"[OrekaShop][fetch_data] Hoan tat: {len(items)} san pham",
        "info",
    )
    return json.dumps(items)


def process_data(data_json, config_json):
    """Chuan hoa du lieu san pham."""
    data = json.loads(data_json) if isinstance(data_json, str) else data_json
    config = json.loads(config_json) if isinstance(config_json, str) else config_json

    total = len(data)
    crawlflow.log(f"[OrekaShop][process] Bat dau chuan hoa {total} san pham", "info")

    normalized = []
    for idx, item in enumerate(data, 1):
        norm = {
            "url": item.get("url", ""),
            "name": (item.get("name") or item.get("product_name") or "").strip(),
            "price": _safe_float(item.get("price", 0)),
            "old_price": _safe_float(item.get("old_price", 0)),
            "image": item.get("image", ""),
            "images": item.get("images"),
            "sku": item.get("sku", "").strip(),
            "description": (item.get("description", "") or "").strip(),
            "category": _clean_category(item.get("category", "")),
            "availability": item.get("availability", "Còn hàng"),
            "crawled_at": item.get("crawled_at", datetime.now().strftime("%Y-%m-%dT%H:%M:%S")),
            "specs": item.get("specs", {}),
        }
        name = norm["name"] or "(khong ten)"
        price_str = f" | gia: {norm['price']}" if norm["price"] else ""
        crawlflow.log(
            f"[OrekaShop][process] [{idx}/{total}] \"{name}\"{price_str} | {norm['url']}",
            "debug",
        )
        normalized.append(norm)

    crawlflow.log(f"[OrekaShop][process] Hoan tat chuan hoa {len(normalized)}/{total} san pham", "info")
    return json.dumps(normalized)


def export_data(data_json, config_json):
    """Xuat du lieu ra Excel/CSV voi co che append + check trung.

    Dinh dang cot: STT, Gia bia, Gia ban, Ton kho, Don vi,
                   Khoi luong (g), Kich thuoc, Tinh trang, Nam XB,
                   Thuong hieu, Ten sach, Danh muc

    Ghi log vao file dedup de tranh trung lap.
    Luon append vao file Excel (doc file cu, them rows moi, ghi de).
    """
    data = json.loads(data_json) if isinstance(data_json, str) else data_json
    config = json.loads(config_json) if isinstance(config_json, str) else config_json

    project_id = config.get("project_id", "default")
    output_dir = config.get("output_dir")
    if not output_dir:
        # Mac dinh dung thu muc Downloads cua user hien tai
        output_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(output_dir, exist_ok=True)

    # Lay ten shop: uu tien projectName tu config (neu co), sau do lay tu URL
    project_name = config.get("projectName", "")
    shop_url = config.get("shop_url", "")
    shop_name = "oreka_shop"
    
    if project_name:
        # Dung project name neu duoc cung cap
        shop_name = project_name.replace("?", "_").replace("&", "_").replace("/", "_").replace("\\", "_")
    elif shop_url:
        # Neu khong co project name, lay tu URL
        parts = shop_url.rstrip("/").split("/")
        if parts:
            shop_name = parts[-1].replace("?", "_").replace("&", "_")

    started_at = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    total_items = len(data)

    crawlflow.log(f"[OrekaShop] Bat dau export {total_items} san pham", "info")

    # ── Dedup: doc file dedup ────────────────────────────────────────
    dedup_path = os.path.join(output_dir, f".{shop_name}_dedup.json")
    seen_ids = set()
    if os.path.exists(dedup_path):
        try:
            content = crawlflow.read_file(dedup_path)
            seen_ids = set(json.loads(content))
            crawlflow.log(f"[OrekaShop] Da doc {len(seen_ids)} ID da xu ly tu file dedup", "info")
        except Exception as e:
            crawlflow.log(f"[OrekaShop] Loi doc dedup file: {e}", "warn")

    # Loc san pham moi
    new_products = []
    for item in data:
        dedup_key = item.get("url", "") or item.get("sku", "")
        if dedup_key and dedup_key in seen_ids:
            continue
        new_products.append(item)

    crawlflow.log(f"[OrekaShop] Sau dedup: {len(new_products)} san pham moi (da bo qua {len(data) - len(new_products)} san pham trung)", "info")

    if not new_products:
        crawlflow.log("[OrekaShop] Khong co san pham moi de export", "info")
        _update_progress(project_id, {
            "items_total": total_items,
            "items_processed": total_items,
            "items_success": 0,
            "items_failed": 0,
            "progress_pct": 100.0,
            "avg_time_ms": 0,
            "total_time_ms": 0,
            "started_at": started_at,
            "message": "Khong co san pham moi",
        })
        return json.dumps({"file": "", "count": 0, "new": 0, "skipped": len(data)})

    # ── Xuat Excel ──────────────────────────────────────────────────
    excel_path = os.path.join(output_dir, f"{shop_name}_products.xlsx")
    csv_path = os.path.join(output_dir, f"{shop_name}_products.csv")

    try:
        from openpyxl import Workbook, load_workbook
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False

    if has_openpyxl:
        count = _export_xlsx(new_products, excel_path, seen_ids)
        crawlflow.log(f"[OrekaShop] Da ghi {count} san pham vao {excel_path}", "info")
    else:
        count = _export_csv(new_products, csv_path)
        crawlflow.log(f"[OrekaShop] Da ghi {count} san pham vao {csv_path} (CSV)", "info")

    # Luu dedup
    for item in new_products:
        dedup_key = item.get("url", "") or item.get("sku", "")
        if dedup_key:
            seen_ids.add(dedup_key)

    try:
        crawlflow.save_file(dedup_path, json.dumps(list(seen_ids), ensure_ascii=False))
    except Exception as e:
        crawlflow.log(f"[OrekaShop] Loi ghi dedup file: {e}", "warn")

    elapsed = (datetime.now() - datetime.strptime(started_at, "%Y-%m-%dT%H:%M:%S")).total_seconds() * 1000
    _update_progress(project_id, {
        "items_total": total_items,
        "items_processed": total_items,
        "items_success": count,
        "items_failed": total_items - count,
        "progress_pct": 100.0,
        "avg_time_ms": elapsed / max(total_items, 1),
        "total_time_ms": elapsed,
        "started_at": started_at,
        "message": f"Export hoan thanh: {count} san pham moi",
    })

    output_file = excel_path if has_openpyxl else csv_path
    result = {
        "file": output_file,
        "count": count,
        "new": len(new_products),
        "skipped": len(data) - len(new_products),
    }
    return json.dumps(result)


def _export_xlsx(products, filepath, seen_ids):
    """Ghi san pham vao file Excel (append neu file da ton tai)."""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        crawlflow.log(f"[OrekaShop] Mo file Excel co san: {filepath}", "info")
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "San pham"
        headers = [
            "STT", "Giá bìa", "Giá bán", "Tồn kho", "Đơn vị",
            "Khối lượng (g)", "Kích thước", "Tình trạng", "Năm XB",
            "Thương hiệu", "Tên sách", "Danh mục"
        ]
        ws.append(headers)
        crawlflow.log(f"[OrekaShop] Tao file Excel moi: {filepath}", "info")

    count = 0

    for item in products:
        dedup_key = item.get("url", "") or item.get("sku", "")
        if dedup_key and dedup_key in seen_ids:
            continue

        specs = item.get("specs", {})

        row = [
            ws.max_row,
            item.get("old_price", 0),
            item.get("price", 0),
            item.get("stock", ""),
            _spec_value(specs, "Đơn vị", "Đơn vị tính"),
            _spec_value(specs, "Trọng lượng", "Khối lượng", "Khối lượng (g)"),
            _spec_value(specs, "Kích thước", "Kích thước sản phẩm"),
            item.get("availability", ""),
            _spec_value(specs, "Năm xuất bản", "Năm XB"),
            _spec_value(specs, "Thương hiệu", "Nhà xuất bản", "NXB", "Tác giả"),
            item.get("name") or item.get("product_name", ""),
            _clean_category(item.get("category", "")),
        ]
        ws.append(row)
        count += 1

    wb.save(filepath)
    crawlflow.log(f"[OrekaShop] Da them {count} dong vao Excel", "info")
    return count


def _export_csv(products, filepath):
    """Fallback: ghi CSV."""
    import csv
    import io

    mode = "a" if os.path.exists(filepath) else "w"
    has_header = mode == "w"

    # Count existing data rows to continue STT
    existing_count = 0
    if mode == "a" and os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8-sig") as f:
            existing_count = sum(1 for _ in f) - 1  # subtract header row

    with open(filepath, mode, newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if has_header:
            writer.writerow([
                "STT", "Giá bìa", "Giá bán", "Tồn kho", "Đơn vị",
                "Khối lượng (g)", "Kích thước", "Tình trạng", "Năm XB",
                "Thương hiệu", "Tên sách", "Danh mục"
            ])

        count = existing_count
        for item in products:
            specs = item.get("specs", {})

            writer.writerow([
                count + 1,
                item.get("old_price", 0),
                item.get("price", 0),
                item.get("stock", ""),
                _spec_value(specs, "Đơn vị", "Đơn vị tính"),
                _spec_value(specs, "Trọng lượng", "Khối lượng", "Khối lượng (g)"),
                _spec_value(specs, "Kích thước", "Kích thước sản phẩm"),
                item.get("availability", ""),
                _spec_value(specs, "Năm xuất bản", "Năm XB"),
                _spec_value(specs, "Thương hiệu", "Nhà xuất bản", "NXB", "Tác giả"),
                item.get("name") or item.get("product_name", ""),
                _clean_category(item.get("category", "")),
            ])
            count += 1

    crawlflow.log(f"[OrekaShop] Da them {count} dong vao CSV", "info")
    return count


def _update_progress(project_id, data):
    """Cap nhat progress vao Rust backend."""
    try:
        def _to_u64(v):
            try:
                return int(round(float(v))) if v is not None else 0
            except (TypeError, ValueError):
                return 0

        info = {
            "items_total": _to_u64(data.get("items_total", 0)),
            "items_processed": _to_u64(data.get("items_processed", 0)),
            "items_success": _to_u64(data.get("items_success", 0)),
            "items_failed": _to_u64(data.get("items_failed", 0)),
            "progress_pct": float(data.get("progress_pct", 0.0) or 0.0),
            "avg_time_ms": float(data.get("avg_time_ms", 0.0) or 0.0),
            "total_time_ms": _to_u64(data.get("total_time_ms", 0)),
            "started_at": data.get("started_at", ""),
            "message": data.get("message", ""),
        }
        crawlflow.update_progress(project_id, json.dumps(info))
    except Exception as e:
        crawlflow.log(f"[OrekaShop] Loi update progress: {e}", "error")




def on_unload():
    crawlflow.log("[OrekaShop] Plugin unloaded", "info")
